import csv
from math import log10

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter


class Vacancy:
    currency_to_ruble = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055}

    def __init__(self, **kwargs):
        self.name = kwargs['name']
        self.area_name = kwargs['area_name']
        spliten = kwargs['published_at'].split('T')
        date = spliten[0].split('-')
        self.year = int(date[0])

        self.salary_currency = kwargs['salary_currency']
        self.salary = int((float(kwargs['salary_from']) + float(kwargs['salary_to'])) // 2 *
                          self.currency_to_ruble[self.salary_currency])


class DataSet:
    def __init__(self, header):
        self.header = header
        self.vacancies_objects = []

    def get_stat(self, vacancy_name):
        vacancies = self.vacancies_objects
        doly_stat = dict()
        salary_stat = dict()
        vacancy_count_stat = dict()
        selected_salary_stat = dict()
        selected_count_stat = dict()
        area_salary_stat = dict()
        area_count_stat = dict()
        for vacancy in vacancies:
            salary = vacancy.salary
            if vacancy.area_name not in doly_stat:
                doly_stat[vacancy.area_name] = 0
                area_salary_stat[vacancy.area_name] = 0
                area_count_stat[vacancy.area_name] = 0
            doly_stat[vacancy.area_name] += 1
            area_salary_stat[vacancy.area_name] += salary
            area_count_stat[vacancy.area_name] += 1

            if vacancy.year not in salary_stat:
                salary_stat[vacancy.year] = 0
                vacancy_count_stat[vacancy.year] = 0
                selected_salary_stat[vacancy.year] = 0
                selected_count_stat[vacancy.year] = 0
            salary_stat[vacancy.year] += salary
            vacancy_count_stat[vacancy.year] += 1
            if vacancy_name in vacancy.name:
                selected_salary_stat[vacancy.year] += salary
                selected_count_stat[vacancy.year] += 1

        doly_stat = {k: doly_stat[k] / len(vacancies) for k in doly_stat if doly_stat[k] >= int(len(vacancies) / 100)}
        doly_stat = {k: round(doly_stat[k], 4) for k in sorted(doly_stat, key=lambda k: -doly_stat[k])}
        salary_stat = {k: salary_stat[k] // vacancy_count_stat[k] for k in sorted(salary_stat)}
        selected_salary_stat = {
            k: selected_salary_stat[k] // selected_count_stat[k] if selected_count_stat[k] != 0 else 0 for k in
            sorted(selected_salary_stat)}
        area_salary_stat = {k: area_salary_stat[k] // area_count_stat[k] for k in area_salary_stat if k in doly_stat}
        area_salary_stat = {k: area_salary_stat[k] for k in
                            sorted(area_salary_stat, key=lambda k: -area_salary_stat[k])}

        report.generate_excel(
            [salary_stat, selected_salary_stat, vacancy_count_stat, selected_count_stat],
            [{k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)},
             {k: doly_stat[k] for i, k in zip(range(10), doly_stat)}]
        )

        """
        print('Динамика уровня зарплат по годам:', salary_stat)
        print('Динамика количества вакансий по годам:', vacancy_count_stat)
        print('Динамика уровня зарплат по годам для выбранной профессии:', selected_salary_stat)
        print('Динамика количества вакансий по годам для выбранной профессии:', selected_count_stat)
        print('Уровень зарплат по городам (в порядке убывания):',
              {k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)})
        print('Доля вакансий по городам (в порядке убывания):', {k: doly_stat[k] for i, k in zip(range(10), doly_stat)})
        """
        return


class Report:
    wb: Workbook
    ws1: Worksheet
    ws2: Worksheet
    thin_border: Border

    def __init__(self, columns1):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Статистика по годам"
        self.ws2 = self.wb.create_sheet("Статистика по городам")

        font = Font(bold=True)
        self.thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for i, c in enumerate(columns1, 1):
            cell = self.ws1.cell(1, i, c)
            cell.font = font
            cell.border = self.thin_border
            self.ws1.column_dimensions[get_column_letter(i)].width = max(len(c)+2, 6)

        cell = self.ws2.cell(1, 1, "Город")
        cell.font = font
        cell.border = self.thin_border
        cell = self.ws2.cell(1, 2, "Уровень зарплат")
        cell.font = font
        cell.border = self.thin_border

        self.ws2.column_dimensions[get_column_letter(3)].width = 2

        cell = self.ws2.cell(1, 4, "Город")
        cell.font = font
        cell.border = self.thin_border
        cell = self.ws2.cell(1, 5, "Доля вакансий")
        cell.font = font
        cell.border = self.thin_border

    def generate_excel(self, data: list[dict], data2: list[dict]):
        a = data[0]  # for keys
        for i, key in enumerate(a, 2):
            year_cell = self.ws1.cell(i, 1, key)
            year_cell.border = self.thin_border
            for j, values in enumerate(data, 2):
                cell = self.ws1.cell(i, j, values[key])
                cell.border = self.thin_border

        width1_max = len(self.ws2.cell(1, 1).value)
        width2_max = len(self.ws2.cell(1, 2).value)
        for i, key in enumerate(data2[0], 2):
            cell = self.ws2.cell(i, 1, key)
            cell.border = self.thin_border
            if len(key) > width1_max:
                width1_max = len(key)
            cell = self.ws2.cell(i, 2, data2[0][key])
            cell.border = self.thin_border
            v = int(log10(data2[0][key]))
            if v > width2_max:
                width2_max = v

        self.ws2.column_dimensions[get_column_letter(1)].width = width1_max + 2
        self.ws2.column_dimensions[get_column_letter(2)].width = width2_max + 2

        width1_max = len(self.ws2.cell(1, 4).value)
        width2_max = len(self.ws2.cell(1, 5).value)
        for i, key in enumerate(data2[1], 2):
            cell = self.ws2.cell(i, 4, key)
            cell.border = self.thin_border
            if len(key) > width1_max:
                width1_max = len(key)
            s = f"{data2[1][key]*100:.2f}%"
            cell = self.ws2.cell(i, 5, s)
            cell.number_format = "0.00%"
            cell.border = self.thin_border
            v = len(s)
            if v > width2_max:
                width2_max = v

        self.ws2.column_dimensions[get_column_letter(4)].width = width1_max + 2
        self.ws2.column_dimensions[get_column_letter(5)].width = width2_max + 2

        self.wb.save("report.xlsx")


def csv_read(filename):
    with open(filename, encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        header = reader.__next__()
        ds = DataSet(header)
        ds.vacancies_objects = [Vacancy(
            **{k: v for k, v in zip(header, line)}) for line in reader if len(line) == len(header) and all(line)]
    return ds


"""
vacancies_by_year.csv
Программист
"""

file_name = input('Введите название файла: ')
vacancy_name = input('Введите название профессии: ')
report = Report([
    "Год",
    "Средняя зарплата",
    "Средняя зарплата - " + vacancy_name,
    "Количество вакансий",
    "Количество вакансий - " + vacancy_name])
data_set = csv_read(file_name)
data_set.get_stat(vacancy_name)
