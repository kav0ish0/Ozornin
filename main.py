import csv
from math import log10

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

import matplotlib.pyplot as plt
import numpy as np

import pdfkit
import os
from jinja2 import Template


def getpath():
    """Возвращает путь к графику

    Returns:
        str: Путь к графику
    """
    return os.path.join(os.path.abspath("."), "graph.png").replace("\\", '/')


def get_percent(v):
    """Конвертирует число от 0 до 1 в процентный вид

    Args:
        v (float): Число, процентный вид которого нужно получить
    Returns:
        str: Процентный вид числа v
    """
    return f"{v * 100:.2f}%"


class Vacancy:
    """Класс для хранения данных о вакансии

    Attributes:
        name (str): Название вакансии
        area_name (str): Название города вакансии
        year (int): Год, в котором были сохранены данные о вакансии
        salary_currency (str): Валюта вакансии
        salary (int): Средняя зарплата вакансии в рублях
    """

    currency_to_ruble = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055}

    def __init__(self, **kwargs):
        """Инициализирует объект Vacancy, выполняет конвертацию для целочисленного поля year

        Named args:
            name (str): Название вакансии
            area_name (str): Название города вакансии
            published_at (int): Дата появления вакансии
            salary_currency (str): Валюта вакансии
            salary_from (int or float or str): Нижняя граница оклада вакансии
            salary_to (int or float or str): Верхняя граница оклада вакансии
        """
        self.name = kwargs['name']
        self.area_name = kwargs['area_name']
        spliten = kwargs['published_at'].split('T')
        date = spliten[0].split('-')
        self.year = int(date[0])

        self.salary_currency = kwargs['salary_currency']
        self.salary = int((float(kwargs['salary_from']) + float(kwargs['salary_to'])) // 2 *
                          self.currency_to_ruble[self.salary_currency])


class DataSet:
    """Класс для хранения данных о всех вакансиях и выводе информации о них

    Attributes:
        header (str[]): Названия полей о вакансии из csv файла
        vacancies_objects (Vacancy[]): Массив с данными о всех вакансиях из csv файла
    """

    def __init__(self, header):
        """Инициализирует объект DataSet"""
        self.header = header
        self.vacancies_objects = []

    def get_stat(self, vacancy_name, print_type):
        """Собирает статистику и данные о вакансиях и просит класс Report вывести их

        Attributes:
            vacancy_name (str): Название вакансии, о которой нужно отдельно собрать статистику
            print_type (int):
                Метод собирает данные о вакансиях в файл Excel, если 0.
                Метод собирает статистику в .pdf файл, если 1.
        """
        vacancies = self.vacancies_objects
        doly_stat = dict()
        salary_stat = dict()
        vacancy_count_stat = dict()
        selected_salary_stat = dict()
        selected_count_stat = dict()
        area_salary_stat = dict()
        area_count_stat = dict()
        for vacancy in vacancies:
            salary = vacancy.salary
            if vacancy.area_name not in doly_stat:
                doly_stat[vacancy.area_name] = 0
                area_salary_stat[vacancy.area_name] = 0
                area_count_stat[vacancy.area_name] = 0
            doly_stat[vacancy.area_name] += 1
            area_salary_stat[vacancy.area_name] += salary
            area_count_stat[vacancy.area_name] += 1

            if vacancy.year not in salary_stat:
                salary_stat[vacancy.year] = 0
                vacancy_count_stat[vacancy.year] = 0
                selected_salary_stat[vacancy.year] = 0
                selected_count_stat[vacancy.year] = 0
            salary_stat[vacancy.year] += salary
            vacancy_count_stat[vacancy.year] += 1
            if vacancy_name in vacancy.name:
                selected_salary_stat[vacancy.year] += salary
                selected_count_stat[vacancy.year] += 1

        doly_stat = {k: doly_stat[k] / len(vacancies) for k in doly_stat if doly_stat[k] >= int(len(vacancies) / 100)}
        doly_stat = {k: round(doly_stat[k], 4) for k in sorted(doly_stat, key=lambda k: -doly_stat[k])}
        salary_stat = {k: salary_stat[k] // vacancy_count_stat[k] for k in sorted(salary_stat)}
        selected_salary_stat = {
            k: selected_salary_stat[k] // selected_count_stat[k] if selected_count_stat[k] != 0 else 0 for k in
            sorted(selected_salary_stat)}
        area_salary_stat = {k: area_salary_stat[k] // area_count_stat[k] for k in area_salary_stat if k in doly_stat}
        area_salary_stat = {k: area_salary_stat[k] for k in
                            sorted(area_salary_stat, key=lambda k: -area_salary_stat[k])}

        if print_type == 0:
            report.generate_excel(
                [salary_stat, selected_salary_stat, vacancy_count_stat, selected_count_stat],
                [{k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)},
                 {k: doly_stat[k] for i, k in zip(range(10), doly_stat)}]
            )
        else:
            report.generate_pdf(salary_stat, vacancy_count_stat, selected_salary_stat,
                                selected_count_stat, area_salary_stat, doly_stat)

        """
        print('Динамика уровня зарплат по годам:', salary_stat)
        print('Динамика количества вакансий по годам:', vacancy_count_stat)
        print('Динамика уровня зарплат по годам для выбранной профессии:', selected_salary_stat)
        print('Динамика количества вакансий по годам для выбранной профессии:', selected_count_stat)
        print('Уровень зарплат по городам (в порядке убывания):',
              {k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)})
        print('Доля вакансий по городам (в порядке убывания):', {k: doly_stat[k] for i, k in zip(range(10), doly_stat)})
        """
        return


class Report:
    """Класс для вывода данных из класса DataSet

    Attributes:
        wb (Workbook): Главный объект хранения данных об Excel файле
        ws1 (Worksheet): Объект страницы в Excel файле для статистики по годам
        ws2 (Worksheet): Объект страницы в Excel файле для статистики по городам
        thin_border (Border): Стиль границы для ячейки в Excel файле
        columns (str[]): Названия колонок для статистики по годам
    """
    wb: Workbook
    ws1: Worksheet
    ws2: Worksheet
    thin_border: Border
    columns = []

    def __init__(self, columns1):
        """Инициализирует объект Report, подгатавливает Excel файл для записи данных

        Arguments:
            columns1 (str[]): Названия колонок для статистики по годам
        """
        self.columns = columns1

        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Статистика по годам"
        self.ws2 = self.wb.create_sheet("Статистика по городам")

        font = Font(bold=True)
        self.thin_border = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        for i, c in enumerate(columns1, 1):
            cell = self.ws1.cell(1, i, c)
            cell.font = font
            cell.border = self.thin_border
            self.ws1.column_dimensions[get_column_letter(i)].width = max(len(c) + 2, 6)

        cell = self.ws2.cell(1, 1, "Город")
        cell.font = font
        cell.border = self.thin_border
        cell = self.ws2.cell(1, 2, "Уровень зарплат")
        cell.font = font
        cell.border = self.thin_border

        self.ws2.column_dimensions[get_column_letter(3)].width = 2

        cell = self.ws2.cell(1, 4, "Город")
        cell.font = font
        cell.border = self.thin_border
        cell = self.ws2.cell(1, 5, "Доля вакансий")
        cell.font = font
        cell.border = self.thin_border

    def generate_excel(self, data: list[dict], data2: list[dict]):
        """Генерация Excel файла

        Arguments:
            data (list[dict]): лист данных о вакансиях в следующем порядке:
                1. Динамика уровня зарплат по годам
                2. Динамика уровня зарплат по годам для выбранной профессии
                3. Динамика количества вакансий по годам
                4. Динамика количества вакансий по годам для выбранной профессии
            data2 (list[dict]): лист статистики данных в следующем порядке:
                1. Уровень зарплат по городам (в порядке убывания, первые 10 значений)
                2. Доля вакансий по городам (в порядке убывания, первые 10 значений)
        """

        a = data[0]  # for keys
        for i, key in enumerate(a, 2):
            year_cell = self.ws1.cell(i, 1, key)
            year_cell.border = self.thin_border
            for j, values in enumerate(data, 2):
                cell = self.ws1.cell(i, j, values[key])
                cell.border = self.thin_border

        width1_max = len(self.ws2.cell(1, 1).value)
        width2_max = len(self.ws2.cell(1, 2).value)
        for i, key in enumerate(data2[0], 2):
            cell = self.ws2.cell(i, 1, key)
            cell.border = self.thin_border
            if len(key) > width1_max:
                width1_max = len(key)
            cell = self.ws2.cell(i, 2, data2[0][key])
            cell.border = self.thin_border
            v = int(log10(data2[0][key]))
            if v > width2_max:
                width2_max = v

        self.ws2.column_dimensions[get_column_letter(1)].width = width1_max + 2
        self.ws2.column_dimensions[get_column_letter(2)].width = width2_max + 2

        width1_max = len(self.ws2.cell(1, 4).value)
        width2_max = len(self.ws2.cell(1, 5).value)
        for i, key in enumerate(data2[1], 2):
            cell = self.ws2.cell(i, 4, key)
            cell.border = self.thin_border
            if len(key) > width1_max:
                width1_max = len(key)
            s = f"{data2[1][key] * 100:.2f}%"
            cell = self.ws2.cell(i, 5, s)
            cell.number_format = "0.00%"
            cell.border = self.thin_border
            v = len(s)
            if v > width2_max:
                width2_max = v

        self.ws2.column_dimensions[get_column_letter(4)].width = width1_max + 2
        self.ws2.column_dimensions[get_column_letter(5)].width = width2_max + 2

        self.wb.save("report.xlsx")

    def generate_image(self, salary_stat, vacancy_count_stat, selected_salary_stat,
                       selected_count_stat, area_salary_stat, doly_stat):
        """Генерация графика в файл graph.png

        Arguments:
            salary_stat (dict): Динамика уровня зарплат по годам
            selected_salary_stat (dict): Динамика уровня зарплат по годам для выбранной профессии
            vacancy_count_stat (dict): Динамика количества вакансий по годам
            selected_count_stat (dict): Динамика количества вакансий по годам для выбранной профессии
            area_salary_stat (dict): Уровень зарплат по городам (в порядке убывания)
            doly_stat (dict): Доля вакансий по городам (в порядке убывания)
        """

        plt.rcParams.update({'font.size': 8})
        fig = plt.figure()

        # Graph 1
        x = np.arange(len(salary_stat))
        width = 0.35
        ax = fig.add_subplot(221)
        ax.bar(x - width / 2, salary_stat.values(), width, label="средняя з/п")
        ax.bar(x + width / 2, selected_salary_stat.values(), width, label="з/п " + vacancy_name.lower())
        ax.set_title("Уровень зарплат по годам")
        ax.set_xticks(x, salary_stat.keys(), rotation="vertical")
        ax.legend()
        ax.grid(axis='y')

        # Graph 2
        x = np.arange(len(vacancy_count_stat))
        ax1 = fig.add_subplot(222)
        ax1.bar(x - width / 2, vacancy_count_stat.values(), width, label="Количество вакансий")
        ax1.bar(x + width / 2, selected_count_stat.values(), width, label="Количество вакансий " + vacancy_name.lower())
        ax1.set_title("Количество вакансий по годам")
        ax1.set_xticks(x, vacancy_count_stat.keys(), rotation="vertical")
        ax1.legend()
        ax1.grid(axis='y')

        # Graph 3
        area_salary_stat = {k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)}
        x = np.arange(len(area_salary_stat))
        ax2 = fig.add_subplot(223)
        ax2.barh(x, area_salary_stat.values(), 0.7)
        ax2.set_title("Уровень зарплат по городам")
        ax2.set_yticks(x, [k.replace("-", "-\n") for k in area_salary_stat.keys()])
        ax2.invert_yaxis()
        ax2.grid(axis='x')
        for item in (ax2.get_yticklabels()):
            item.set_fontsize(6)

        # Graph 4
        x = np.arange(len(doly_stat))
        ax3 = fig.add_subplot(224)
        if len(doly_stat) <= 10:
            ax3.pie(doly_stat.values(), labels=doly_stat.keys(), textprops={'fontsize': 6})
        else:
            rest = sum(list(doly_stat.values())[10:])
            ax3.pie(list(doly_stat.values())[:10] + [rest], labels=list(doly_stat.keys())[:10] + ["Другие"],
                    textprops={'fontsize': 6})
        ax3.set_title("Доля зарплат по городам")

        fig.tight_layout()
        plt.savefig("graph.png")

    def generate_pdf(self, salary_stat, vacancy_count_stat, selected_salary_stat,
                     selected_count_stat, area_salary_stat, doly_stat):
        """Генерация статистики в файл report.pdf

        Arguments:
            salary_stat (dict): Динамика уровня зарплат по годам
            selected_salary_stat (dict): Динамика уровня зарплат по годам для выбранной профессии
            vacancy_count_stat (dict): Динамика количества вакансий по годам
            selected_count_stat (dict): Динамика количества вакансий по годам для выбранной профессии
            area_salary_stat (dict): Уровень зарплат по городам (в порядке убывания)
            doly_stat (dict): Доля вакансий по городам (в порядке убывания)
        """

        report.generate_image(salary_stat, vacancy_count_stat, selected_salary_stat,
                              selected_count_stat, area_salary_stat, doly_stat)

        data = [salary_stat, selected_salary_stat, vacancy_count_stat, selected_count_stat]
        data2 = [{k: area_salary_stat[k] for i, k in zip(range(10), area_salary_stat)},
                 {k: doly_stat[k] for i, k in zip(range(10), doly_stat)}]
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        template = Template(r"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {
                  font-family: Verdana, sans-serif;
                }
            </style>
        </head>
        <body>
        <center>

            <h1>
        Аналитика по зарплатам и городам для профессии {{ vacancy_name }}
            </h1>

        <img src="file:///{{ getpath() }}">

        <h2>Статистика по годам</h2>
        <table border="1" cellpadding="5">
            <tr>
            {% for c in columns %}
                <th>{{ c }}</th>
            {% endfor %}
            </tr>
            {% for year in data[0] %}
            <tr>
                <td><center>{{ year }}</center></td>
                {% for d in data %}
                <td><center>{{ d[year] }}</center></td>
                {% endfor %}
            </tr>
            {% endfor %}
        </table>

        <h2>Статистика по городам</h2>
        <table>
            <tr>
                <td>
                    <table border="1" cellpadding="5">
                        <tr>
                            <th>Город</th>
                            <th>Уровень зарплат</th>
                        </tr>
                        {% for city in data2[0] %}
                        <tr>
                            <td><center>{{city}}</center></td>
                            <td><center>{{data2[0][city]}}</center></td>
                        </tr>
                        {% endfor %}
                    </table>
                </td>
                <td>
                    <table border="1" cellpadding="5">
                        <tr>
                            <th>Город</th>
                            <th>Доля вакансий</th>
                        </tr>
                        {% for city in data2[1] %}
                        <tr>
                            <td><center>{{city}}</center></td>
                            <td><center>{{get_percent(data2[1][city])}}</center></td>
                        </tr>
                        {% endfor %}
                    </table>
                </td>
            </tr>
        </table>

        </center>
        </body>
        </html>
        """)

        pdfkit.from_string(
            template.render(getpath=getpath, get_percent=get_percent, vacancy_name=vacancy_name, columns=self.columns,
                            textstart='<center><p style="font-family: Verdana">', textend="</p></center>",
                            data=data, data2=data2),
            'report.pdf', configuration=config, options={"enable-local-file-access": ""})


def csv_read(filename):
    """Считывание данных о вакансиях из .csv файла

    Arguments:
        filename (str): Путь к файлу .csv с данными о вакансиях
    Returns:
        DataSet: Объект DataSet
    """
    with open(filename, encoding='utf-8-sig') as file:
        reader = csv.reader(file)
        header = reader.__next__()
        ds = DataSet(header)
        ds.vacancies_objects = [Vacancy(
            **{k: v for k, v in zip(header, line)}) for line in reader if len(line) == len(header) and all(line)]
    return ds


"""
vacancies_by_year.csv
Программист
"""

file_name = input('Введите название файла: ')
vacancy_name = input('Введите название профессии: ')
printing_type = input('Вакансии или Статистика?: ')
if printing_type == "Вакансии":
    printing_type = 0
else:
    printing_type = 1
report = Report([
    "Год",
    "Средняя зарплата",
    "Средняя зарплата - " + vacancy_name,
    "Количество вакансий",
    "Количество вакансий - " + vacancy_name])
data_set = csv_read(file_name)
data_set.get_stat(vacancy_name, printing_type)
